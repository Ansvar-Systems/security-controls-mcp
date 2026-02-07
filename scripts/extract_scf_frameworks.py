#!/usr/bin/env python3
"""
Extract all frameworks from SCF 2025.4 spreadsheet.

This script parses the official SCF Excel file and generates:
1. scf-controls.json - All controls with complete framework mappings
2. framework-to-scf.json - Reverse index from framework controls to SCF IDs

Usage:
    poetry run python scripts/extract_scf_frameworks.py
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
SCF_XLSX = SCRIPT_DIR / "data" / "scf-2025.4.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "src" / "security_controls_mcp" / "data"

# Framework column mapping: column index -> (framework_id, display_name)
# These are extracted from the SCF 2025.4 spreadsheet headers
FRAMEWORK_COLUMNS = {
    # SOC 2 / AICPA
    24: ("soc_2_tsc", "AICPA TSC 2017:2022 (SOC 2)"),

    # Privacy Frameworks
    25: ("apec_privacy_2015", "APEC Privacy Framework 2015"),
    36: ("gapp", "Generally Accepted Privacy Principles (GAPP)"),
    93: ("oecd_privacy", "OECD Privacy Principles"),

    # German Standards
    26: ("bsi_200_1", "BSI Standard 200-1"),

    # CIS Controls
    27: ("cis_csc_8.1", "CIS Critical Security Controls v8.1"),
    28: ("cis_csc_8.1_ig1", "CIS CSC v8.1 Implementation Group 1"),
    29: ("cis_csc_8.1_ig2", "CIS CSC v8.1 Implementation Group 2"),
    30: ("cis_csc_8.1_ig3", "CIS CSC v8.1 Implementation Group 3"),

    # Governance Frameworks
    31: ("cobit_2019", "COBIT 2019"),
    32: ("coso_2017", "COSO 2017"),

    # Cloud Security
    33: ("csa_ccm_4", "CSA Cloud Controls Matrix v4"),
    34: ("csa_iot_scf_2", "CSA IoT Security Controls Framework 2"),

    # ENISA
    35: ("enisa_2.0", "ENISA 2.0"),

    # GovRAMP (StateRAMP predecessor)
    37: ("govramp_core", "GovRAMP Core"),
    38: ("govramp_low", "GovRAMP Low"),
    39: ("govramp_low_plus", "GovRAMP Low+"),
    40: ("govramp_moderate", "GovRAMP Moderate"),
    41: ("govramp_high", "GovRAMP High"),

    # IEC Standards (Industrial/Medical)
    42: ("iec_tr_60601_4_5_2021", "IEC TR 60601-4-5:2021 (Medical IT)"),
    43: ("iec_62443_4_2_2019", "IEC 62443-4-2:2019 (Industrial Security)"),

    # Maritime
    44: ("imo_maritime_cyber", "IMO Maritime Cyber Risk Management"),

    # ISO Standards
    45: ("iso_sae_21434_2021", "ISO/SAE 21434:2021 (Automotive Cybersecurity)"),
    46: ("iso_22301_2019", "ISO/IEC 22301:2019 (Business Continuity)"),
    47: ("iso_27001_2022", "ISO/IEC 27001:2022"),
    48: ("iso_27002_2022", "ISO/IEC 27002:2022"),
    49: ("iso_27017_2015", "ISO/IEC 27017:2015 (Cloud Security)"),
    50: ("iso_27018_2014", "ISO/IEC 27018:2014 (Cloud Privacy)"),
    51: ("iso_27701_2025", "ISO/IEC 27701:2025 (Privacy Extension)"),
    52: ("iso_29100_2024", "ISO/IEC 29100:2024 (Privacy Framework)"),
    53: ("iso_31000_2009", "ISO 31000:2009 (Risk Management)"),
    54: ("iso_31010_2009", "ISO 31010:2009 (Risk Assessment)"),
    55: ("iso_42001_2023", "ISO/IEC 42001:2023 (AI Management System)"),

    # MITRE
    56: ("mitre_attack_10", "MITRE ATT&CK v10"),

    # Media/Entertainment
    57: ("mpa_csp_5.1", "MPA Content Security Program 5.1"),

    # Insurance
    58: ("naic_mdl_668", "NAIC Insurance Data Security Model Law (MDL-668)"),

    # NIST AI Frameworks
    59: ("nist_ai_rmf_1.0", "NIST AI 100-1 (AI Risk Management Framework) 1.0"),
    60: ("nist_ai_600_1", "NIST AI 600-1 (Generative AI Profile)"),

    # NIST Privacy
    61: ("nist_privacy_framework_1.0", "NIST Privacy Framework 1.0"),

    # NIST 800 Series
    62: ("nist_800_37_r2", "NIST SP 800-37 R2 (Risk Management Framework)"),
    63: ("nist_800_39", "NIST SP 800-39 (Risk Management)"),
    64: ("nist_800_53_r4", "NIST SP 800-53 R4"),
    65: ("nist_800_53_r4_low", "NIST SP 800-53 R4 (Low)"),
    66: ("nist_800_53_r4_moderate", "NIST SP 800-53 R4 (Moderate)"),
    67: ("nist_800_53_r4_high", "NIST SP 800-53 R4 (High)"),
    68: ("nist_800_53_r5", "NIST SP 800-53 R5"),
    69: ("nist_800_53b_r5_privacy", "NIST SP 800-53B R5 (Privacy)"),
    70: ("nist_800_53b_r5_low", "NIST SP 800-53B R5 (Low)"),
    71: ("nist_800_53b_r5_moderate", "NIST SP 800-53B R5 (Moderate)"),
    72: ("nist_800_53b_r5_high", "NIST SP 800-53B R5 (High)"),
    73: ("nist_800_53_r5_noc", "NIST SP 800-53 R5 (NOC)"),
    74: ("nist_800_63b", "NIST SP 800-63B (Digital Identity)"),
    75: ("nist_800_82_r3_low", "NIST SP 800-82 R3 OT Overlay (Low)"),
    76: ("nist_800_82_r3_moderate", "NIST SP 800-82 R3 OT Overlay (Moderate)"),
    77: ("nist_800_82_r3_high", "NIST SP 800-82 R3 OT Overlay (High)"),
    78: ("nist_800_160", "NIST SP 800-160 (Systems Security Engineering)"),
    79: ("nist_800_161_r1", "NIST SP 800-161 R1 (Supply Chain)"),
    80: ("nist_800_161_r1_baseline", "NIST SP 800-161 R1 C-SCRM Baseline"),
    81: ("nist_800_161_r1_flowdown", "NIST SP 800-161 R1 Flow Down"),
    82: ("nist_800_161_r1_level1", "NIST SP 800-161 R1 Level 1"),
    83: ("nist_800_161_r1_level2", "NIST SP 800-161 R1 Level 2"),
    84: ("nist_800_161_r1_level3", "NIST SP 800-161 R1 Level 3"),
    85: ("nist_800_171_r2", "NIST SP 800-171 R2 (CUI)"),
    86: ("nist_800_171a", "NIST SP 800-171A (Assessment)"),
    87: ("nist_800_171_r3", "NIST SP 800-171 R3 (CUI)"),
    88: ("nist_800_171a_r3", "NIST SP 800-171A R3 (Assessment)"),
    89: ("nist_800_172", "NIST SP 800-172 (Enhanced CUI)"),
    90: ("nist_800_207", "NIST SP 800-207 (Zero Trust)"),
    91: ("nist_800_218", "NIST SP 800-218 (SSDF)"),
    92: ("nist_csf_2.0", "NIST Cybersecurity Framework 2.0"),

    # OWASP
    94: ("owasp_top_10_2021", "OWASP Top 10 2021"),

    # PCI DSS
    95: ("pci_dss_4.0.1", "PCI DSS v4.0.1"),
    96: ("pci_dss_4.0.1_saq_a", "PCI DSS v4.0.1 SAQ A"),
    97: ("pci_dss_4.0.1_saq_a_ep", "PCI DSS v4.0.1 SAQ A-EP"),
    98: ("pci_dss_4.0.1_saq_b", "PCI DSS v4.0.1 SAQ B"),
    99: ("pci_dss_4.0.1_saq_b_ip", "PCI DSS v4.0.1 SAQ B-IP"),
    100: ("pci_dss_4.0.1_saq_c", "PCI DSS v4.0.1 SAQ C"),
    101: ("pci_dss_4.0.1_saq_c_vt", "PCI DSS v4.0.1 SAQ C-VT"),
    102: ("pci_dss_4.0.1_saq_d_merchant", "PCI DSS v4.0.1 SAQ D (Merchant)"),
    103: ("pci_dss_4.0.1_saq_d_sp", "PCI DSS v4.0.1 SAQ D (Service Provider)"),
    104: ("pci_dss_4.0.1_saq_p2pe", "PCI DSS v4.0.1 SAQ P2PE"),

    # Shared Assessments
    105: ("shared_assessments_sig_2025", "Shared Assessments SIG 2025"),

    # Aerospace/Defense
    106: ("sparta", "SPARTA (Space Attack Research)"),

    # SWIFT
    107: ("swift_cscf_2023", "SWIFT Customer Security Framework 2023"),

    # Automotive
    108: ("tisax_isa_6", "TISAX ISA 6 (Automotive)"),
    109: ("ul_2900_1_2017", "UL 2900-1:2017 (Software Cybersecurity)"),
    110: ("un_r155", "UN R155 (Vehicle Cybersecurity)"),
    111: ("un_ece_wp29", "UN ECE WP.29 (Automotive)"),

    # US Frameworks
    112: ("us_c2m2_2.1", "US C2M2 2.1 (Capability Maturity)"),
    113: ("us_cert_rmm_1.2", "US CERT RMM 1.2 (Resilience)"),
    114: ("us_cisa_cpg_2022", "CISA Cross-Sector CPG 2022"),
    115: ("cjis_5.9.3", "CJIS Security Policy v5.9.3"),
    116: ("cmmc_2.0_level_1", "CMMC 2.0 Level 1"),
    117: ("cmmc_2.0_level_1_aos", "CMMC 2.0 Level 1 AOs"),
    118: ("cmmc_2.0_level_2", "CMMC 2.0 Level 2"),
    119: ("cmmc_2.0_level_3", "CMMC 2.0 Level 3"),
    120: ("cms_mars_e_2.0", "CMS MARS-E 2.0 (Healthcare Exchanges)"),
    121: ("us_coppa", "US COPPA (Children's Privacy)"),
    122: ("us_dpf", "US Data Privacy Framework"),
    123: ("dod_zt_roadmap", "DoD Zero Trust Execution Roadmap"),
    124: ("dod_ztra_2.0", "DoD Zero Trust Reference Architecture 2.0"),
    125: ("dfars_252_204_70xx", "DFARS 252.204-70xx (Cybersecurity)"),
    126: ("dhs_cisa_ssdaf", "DHS CISA SSDAF"),
    127: ("dhs_cisa_tic_3.0", "DHS CISA TIC 3.0"),
    128: ("dhs_ztcf", "DHS Zero Trust Capability Framework"),
    129: ("eo_14028", "EO 14028 (Improving Cybersecurity)"),
    130: ("us_facta", "US FACTA"),
    131: ("far_52_204_21", "FAR 52.204-21 (Basic Safeguarding)"),
    132: ("far_52_204_25", "FAR 52.204-25 (NDAA Section 889)"),
    133: ("far_52_204_27", "FAR 52.204-27"),
    134: ("fca_crm", "FCA CRM"),
    135: ("fda_21_cfr_part_11", "FDA 21 CFR Part 11 (Electronic Records)"),
    136: ("fedramp_r4", "FedRAMP R4"),
    137: ("fedramp_r4_low", "FedRAMP R4 (Low)"),
    138: ("fedramp_r4_moderate", "FedRAMP R4 (Moderate)"),
    139: ("fedramp_r4_high", "FedRAMP R4 (High)"),
    140: ("fedramp_r4_lisaas", "FedRAMP R4 (LI-SaaS)"),
    141: ("fedramp_r5", "FedRAMP R5"),
    142: ("fedramp_r5_low", "FedRAMP R5 (Low)"),
    143: ("fedramp_r5_moderate", "FedRAMP R5 (Moderate)"),
    144: ("fedramp_r5_high", "FedRAMP R5 (High)"),
    145: ("fedramp_r5_lisaas", "FedRAMP R5 (LI-SaaS)"),
    146: ("us_ferpa", "US FERPA (Education Privacy)"),
    147: ("ffiec", "FFIEC Cybersecurity Assessment"),
    148: ("us_finra", "US FINRA"),
    149: ("us_fipps", "US FIPPs (Fair Information Practice)"),
    150: ("ftc_act", "FTC Act"),
    151: ("glba_cfr_314_2023", "GLBA CFR 314 (Dec 2023)"),
    152: ("hhs_45_cfr_155_260", "HHS 45 CFR 155.260"),
    153: ("hipaa_admin_2013", "HIPAA Administrative Simplification 2013"),
    154: ("hipaa_security_rule", "HIPAA Security Rule / NIST SP 800-66 R2"),
    155: ("hipaa_hicp_small", "HIPAA HICP Small Practice"),
    156: ("hipaa_hicp_medium", "HIPAA HICP Medium Practice"),
    157: ("hipaa_hicp_large", "HIPAA HICP Large Practice"),
    158: ("irs_1075", "IRS Publication 1075"),
    159: ("itar_part_120", "ITAR Part 120"),
    160: ("nerc_cip_2024", "NERC CIP 2024"),
    161: ("nispom_2020", "NISPOM 2020"),
    162: ("us_nnpi", "US NNPI (Unclassified)"),
    163: ("nstc_nspm_33", "NSTC NSPM-33"),
    164: ("sec_cybersecurity_rule", "SEC Cybersecurity Rule"),
    165: ("sox", "Sarbanes-Oxley Act (SOX)"),
    166: ("ssa_eiesr_8.0", "SSA EIESR 8.0"),
    167: ("tsa_dhs_1580_82_2022", "TSA/DHS 1580/82-2022-01"),

    # US State Laws
    168: ("us_ak_pipa", "Alaska PIPA"),
    169: ("us_ca_sb327", "California SB327 (IoT)"),
    170: ("us_ca_ccpa_2025", "California CCPA/CPRA 2025"),
    171: ("us_ca_sb1386", "California SB1386"),
    172: ("us_co_cpa", "Colorado Privacy Act"),
    173: ("us_il_bipa", "Illinois BIPA (Biometric)"),
    174: ("us_il_ipa", "Illinois IPA"),
    175: ("us_il_pipa", "Illinois PIPA"),
    176: ("us_ma_201_cmr_17", "Massachusetts 201 CMR 17.00"),
    177: ("us_nv_noge_reg_5", "Nevada NOGE Reg 5"),
    178: ("us_nv_sb220", "Nevada SB220"),
    179: ("nydfs_500_2023", "NY DFS 23 NYCRR 500 (2023 Amendment)"),
    180: ("us_ny_shield", "New York SHIELD Act"),
    181: ("us_or_646a", "Oregon 646A"),
    182: ("us_or_cpa", "Oregon Consumer Privacy Act"),
    183: ("us_tn_tipa", "Tennessee TIPA"),
    184: ("us_tx_bc521", "Texas BC521"),
    185: ("us_tx_cdpa", "Texas CDPA"),
    186: ("us_tx_dir_2.0", "Texas DIR Control Standards 2.0"),
    187: ("us_tx_sb820", "Texas SB 820"),
    188: ("us_tx_sb2610", "Texas SB 2610"),
    189: ("tx_ramp_level_1", "TX-RAMP Level 1"),
    190: ("tx_ramp_level_2", "TX-RAMP Level 2"),
    191: ("us_va_cdpa_2025", "Virginia CDPA 2025"),
    192: ("us_vt_act_171", "Vermont Act 171 of 2018"),

    # EMEA - EU Regulations
    193: ("eu_ai_act", "EU AI Act (Regulation 2024/1689)"),
    194: ("eu_cyber_resilience_act", "EU Cyber Resilience Act"),
    195: ("eu_cra_annexes", "EU Cyber Resilience Act Annexes"),
    196: ("eu_eba_gl_2019_04", "EU EBA GL/2019/04"),
    197: ("dora", "Digital Operational Resilience Act (DORA)"),
    198: ("gdpr", "General Data Protection Regulation (GDPR)"),
    199: ("nis2", "NIS2 Directive"),
    200: ("nis2_annex", "NIS2 Directive Annex"),
    201: ("psd2", "PSD2 (Payment Services Directive)"),

    # EMEA - National
    202: ("austria", "Austria Cybersecurity"),
    203: ("belgium", "Belgium Cybersecurity"),
    204: ("germany", "Germany Cybersecurity"),
    205: ("germany_bait", "Germany BAIT (Banking IT)"),
    206: ("germany_c5_2020", "Germany C5:2020 (Cloud)"),
    207: ("greece", "Greece Cybersecurity"),
    208: ("hungary", "Hungary Cybersecurity"),
    209: ("ireland", "Ireland Cybersecurity"),
    210: ("israel_cdmo_1.0", "Israel CDMO 1.0"),
    211: ("israel", "Israel Cybersecurity"),
    212: ("italy", "Italy Cybersecurity"),
    213: ("kenya_dpa_2019", "Kenya DPA 2019"),
    214: ("netherlands", "Netherlands Cybersecurity"),
    215: ("nigeria_dpr_2019", "Nigeria DPR 2019"),
    216: ("norway", "Norway Cybersecurity"),
    217: ("poland", "Poland Cybersecurity"),
    218: ("qatar_pdppl", "Qatar PDPPL"),
    219: ("russia", "Russia Cybersecurity"),
    220: ("saudi_cscc_1_2019", "Saudi Arabia CSCC-1 2019"),
    221: ("saudi_cgiot_1_2024", "Saudi Arabia IoT CGIoT-1 2024"),
    222: ("saudi_ecc_1_2018", "Saudi Arabia ECC-1 2018"),
    223: ("saudi_otcc_1_2022", "Saudi Arabia OTCC-1 2022"),
    224: ("saudi_pdpl", "Saudi Arabia PDPL"),
    225: ("saudi_sacs_002", "Saudi Arabia SACS-002"),
    226: ("saudi_sama_csf_1.0", "Saudi Arabia SAMA CSF 1.0"),
    227: ("serbia_87_2018", "Serbia 87/2018"),
    228: ("south_africa", "South Africa (POPIA)"),
    229: ("spain_boe_a_2022_7191", "Spain BOE-A-2022-7191"),
    230: ("spain_1720_2007", "Spain 1720/2007"),
    231: ("spain_311_2022", "Spain 311/2022"),
    232: ("spain_ccn_stic_825", "Spain CCN-STIC 825"),
    233: ("sweden", "Sweden Cybersecurity"),
    234: ("switzerland", "Switzerland Cybersecurity"),
    235: ("turkey", "Turkey Cybersecurity"),
    236: ("uae_niaf", "UAE NIAF"),
    237: ("uk_caf_4.0", "UK Cyber Assessment Framework 4.0"),
    238: ("uk_cap_1850", "UK CAP 1850"),
    239: ("uk_cyber_essentials", "UK Cyber Essentials"),
    240: ("uk_defstan_05_138", "UK DEFSTAN 05-138"),
    241: ("uk_dpa", "UK Data Protection Act"),

    # APAC
    242: ("australia_essential_8", "Australian Essential Eight"),
    243: ("australia_privacy_act", "Australian Privacy Act"),
    244: ("australia_privacy_principles", "Australian Privacy Principles"),
    245: ("australia_ism_2024", "Australian ISM (June 2024)"),
    246: ("australia_iot_cop", "Australia IoT Code of Practice"),
    247: ("australia_cps_230", "Australia Prudential Standard CPS 230"),
    248: ("australia_cps_234", "Australia Prudential Standard CPS 234"),
    249: ("china_cybersecurity_law", "China Cybersecurity Law"),
    250: ("china_data_security_law", "China Data Security Law"),
    251: ("china_dnsip", "China DNSIP"),
    252: ("china_privacy_law", "China Privacy Law (PIPL)"),
    253: ("hong_kong", "Hong Kong Cybersecurity"),
    254: ("india_dpdpa_2023", "India DPDPA 2023"),
    255: ("india_itr", "India ITR"),
    256: ("india_sebi_cscrf", "India SEBI CSCRF"),
    257: ("japan_appi", "Japan APPI"),
    258: ("japan_ismap", "Japan ISMAP"),
    259: ("malaysia", "Malaysia Cybersecurity"),
    260: ("nz_hisf_2022", "New Zealand HISF 2022"),
    261: ("nz_hisf_suppliers_2023", "New Zealand HISF Suppliers 2023"),
    262: ("nz_nzism_3.6", "New Zealand NZISM 3.6"),
    263: ("nz_privacy_act_2020", "New Zealand Privacy Act 2020"),
    264: ("philippines", "Philippines Cybersecurity"),
    265: ("singapore", "Singapore Cybersecurity"),
    266: ("singapore_cyber_hygiene", "Singapore Cyber Hygiene Practice"),
    267: ("singapore_mas_trm_2021", "Singapore MAS TRM 2021"),
    268: ("south_korea", "South Korea Cybersecurity"),
    269: ("taiwan", "Taiwan Cybersecurity"),

    # Americas (non-US)
    270: ("argentina_ppl", "Argentina PPL"),
    271: ("argentina_reg_132_2018", "Argentina Reg 132-2018"),
    272: ("bahamas", "Bahamas Cybersecurity"),
    273: ("bermuda_bmaccc", "Bermuda BMACCC"),
    274: ("brazil_lgpd", "Brazil LGPD"),
    275: ("canada_csag", "Canada CSAG"),
    276: ("canada_osfi_b13", "Canada OSFI B-13"),
    277: ("canada_itsp_10_171", "Canada ITSP-10-171"),
    278: ("canada_pipeda", "Canada PIPEDA"),
    279: ("chile", "Chile Cybersecurity"),
    280: ("colombia", "Colombia Cybersecurity"),
    281: ("costa_rica", "Costa Rica Cybersecurity"),
    282: ("mexico", "Mexico Cybersecurity"),
    283: ("peru", "Peru Cybersecurity"),
    284: ("uruguay", "Uruguay Cybersecurity"),
}

# Framework categories for organization
FRAMEWORK_CATEGORIES = {
    "ai_governance": [
        "iso_42001_2023", "nist_ai_rmf_1.0", "nist_ai_600_1", "eu_ai_act"
    ],
    "cloud_security": [
        "iso_27017_2015", "iso_27018_2014", "csa_ccm_4", "csa_iot_scf_2",
        "germany_c5_2020"
    ],
    "privacy": [
        "gdpr", "iso_27701_2025", "nist_privacy_framework_1.0", "gapp",
        "apec_privacy_2015", "oecd_privacy", "us_dpf", "brazil_lgpd",
        "us_ca_ccpa_2025", "india_dpdpa_2023", "china_privacy_law"
    ],
    "us_federal": [
        "nist_csf_2.0", "nist_800_53_r5", "fedramp_r5_moderate",
        "fedramp_r5_high", "cmmc_2.0_level_2", "cjis_5.9.3"
    ],
    "financial": [
        "pci_dss_4.0.1", "sox", "glba_cfr_314_2023", "ffiec", "dora", "psd2"
    ],
    "healthcare": [
        "hipaa_security_rule", "hipaa_hicp_small", "hipaa_hicp_medium",
        "hipaa_hicp_large", "cms_mars_e_2.0"
    ],
    "industrial_ot": [
        "iec_62443_4_2_2019", "nerc_cip_2024", "nist_800_82_r3_moderate",
        "nist_800_82_r3_high"
    ],
    "automotive": [
        "iso_sae_21434_2021", "tisax_isa_6", "un_r155", "un_ece_wp29"
    ],
    "supply_chain": [
        "nist_800_161_r1", "nist_800_161_r1_baseline", "dfars_252_204_70xx"
    ],
}


def parse_cell_value(value) -> list[str]:
    """Parse a cell value into a list of control IDs."""
    if not value:
        return []

    # Convert to string and clean
    text = str(value).strip()
    if not text or text.lower() == "none":
        return []

    # Split by newlines and clean each line
    controls = []
    for line in text.split("\n"):
        line = line.strip()
        if line and line.lower() != "none":
            # Handle multiple IDs on same line (comma or semicolon separated)
            for part in re.split(r"[,;]", line):
                part = part.strip()
                if part:
                    controls.append(part)

    return controls


def extract_controls(ws) -> list[dict]:
    """Extract all controls from the worksheet."""
    controls = []

    print("Extracting controls...")
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[2]:  # SCF # column
            continue

        scf_id = str(row[2]).strip()
        if not scf_id or scf_id == "SCF #":
            continue

        # Extract core fields
        control = {
            "id": scf_id,
            "domain": str(row[0]).strip() if row[0] else "",
            "name": str(row[3]).strip() if row[3] else "",
            "description": str(row[3]).strip() if row[3] else "",  # Same as name in SCF
            "weight": int(row[12]) if row[12] and str(row[12]).isdigit() else 5,
            "pptdf": str(row[13]).strip() if row[13] else "",
            "validation_cadence": str(row[4]).strip() if row[4] else "Annual",
            "framework_mappings": {}
        }

        # Extract all framework mappings
        for col_idx, (fw_id, fw_name) in FRAMEWORK_COLUMNS.items():
            try:
                cell_value = row[col_idx] if col_idx < len(row) else None
                mappings = parse_cell_value(cell_value)
                control["framework_mappings"][fw_id] = mappings if mappings else None
            except IndexError:
                control["framework_mappings"][fw_id] = None

        controls.append(control)
        row_count += 1

        if row_count % 100 == 0:
            print(f"  Processed {row_count} controls...")

    print(f"Extracted {len(controls)} controls")
    return controls


def build_reverse_index(controls: list[dict]) -> dict[str, dict[str, list[str]]]:
    """Build reverse index from framework controls to SCF IDs."""
    reverse_index = {}

    for fw_id in FRAMEWORK_COLUMNS.values():
        fw_key = fw_id[0]
        reverse_index[fw_key] = {}

    for control in controls:
        scf_id = control["id"]
        for fw_key, mappings in control["framework_mappings"].items():
            if mappings:
                for mapping in mappings:
                    if mapping not in reverse_index.get(fw_key, {}):
                        reverse_index[fw_key][mapping] = []
                    reverse_index[fw_key][mapping].append(scf_id)

    return reverse_index


def get_framework_stats(controls: list[dict]) -> dict[str, int]:
    """Calculate control counts per framework."""
    stats = {}
    for fw_id in FRAMEWORK_COLUMNS.values():
        fw_key = fw_id[0]
        count = sum(1 for c in controls if c["framework_mappings"].get(fw_key))
        stats[fw_key] = count
    return stats


def main():
    print(f"Loading SCF spreadsheet: {SCF_XLSX}")
    wb = load_workbook(SCF_XLSX, data_only=True)
    ws = wb["SCF 2025.4"]

    # Extract controls
    controls = extract_controls(ws)
    wb.close()

    # Build reverse index
    print("Building reverse index...")
    reverse_index = build_reverse_index(controls)

    # Calculate stats
    stats = get_framework_stats(controls)

    # Output statistics
    print("\n=== FRAMEWORK STATISTICS ===")
    print(f"Total frameworks: {len(FRAMEWORK_COLUMNS)}")
    print(f"Total controls: {len(controls)}")

    # Sort by control count
    sorted_stats = sorted(stats.items(), key=lambda x: x[1], reverse=True)
    print("\nTop 20 frameworks by control count:")
    for fw_id, count in sorted_stats[:20]:
        print(f"  {fw_id}: {count} controls")

    # Tier 0 stats
    print("\n=== TIER 0 (AI GOVERNANCE) ===")
    tier0 = ["iso_42001_2023", "nist_ai_rmf_1.0", "nist_ai_600_1", "eu_ai_act"]
    for fw_id in tier0:
        count = stats.get(fw_id, 0)
        fw_name = next((v[1] for v in FRAMEWORK_COLUMNS.values() if v[0] == fw_id), fw_id)
        print(f"  {fw_name}: {count} controls")

    # Save outputs
    print(f"\nSaving to {OUTPUT_DIR}...")

    # Save controls
    controls_file = OUTPUT_DIR / "scf-controls.json"
    with open(controls_file, "w", encoding="utf-8") as f:
        json.dump({"controls": controls}, f, indent=2, ensure_ascii=False)
    print(f"  Saved {controls_file}")

    # Save reverse index
    reverse_file = OUTPUT_DIR / "framework-to-scf.json"
    with open(reverse_file, "w", encoding="utf-8") as f:
        json.dump(reverse_index, f, indent=2, ensure_ascii=False)
    print(f"  Saved {reverse_file}")

    # Save framework metadata for data_loader.py
    metadata_file = SCRIPT_DIR / "data" / "framework-metadata.json"
    metadata = {
        "frameworks": {v[0]: {"id": v[0], "name": v[1]} for v in FRAMEWORK_COLUMNS.values()},
        "categories": FRAMEWORK_CATEGORIES,
        "statistics": stats,
    }
    with open(metadata_file, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    print(f"  Saved {metadata_file}")

    print("\nDone!")


if __name__ == "__main__":
    main()
